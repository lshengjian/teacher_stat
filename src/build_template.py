# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from src.config import get_options

def build_template(path="专业-工号-姓名-个人档案.xlsx", max_rows=200):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "基本资料"
    ws2 = wb.create_sheet("项目资料")
    dict_ws = wb.create_sheet("数据字典")

    # 数据字典：职称/状态/性别
    titles = get_options("title_categories")  # ["无", "助教", "讲师", "副教授", "教授", "其他高级"]
    statuses = get_options("project_statuses")  # ["申报中", "立项","延期", "结题"]
    genders =  ["男", "女"]

    for i, v in enumerate(titles, 1):
        dict_ws.cell(row=i, column=1, value=v)  # A列：职称
    for i, v in enumerate(statuses, 1):
        dict_ws.cell(row=i, column=2, value=v)  # B列：状态
    for i, v in enumerate(genders, 1):
        dict_ws.cell(row=i, column=3, value=v)  # C列：性别



    # Sheet1：基本资料
    headers1 = ["专业", "工号", "姓名", "职称", "性别", "生日", "入校日期"]
    ws1.append(headers1)
    ws1["A2"] = "='数据字典'!D6"
    ws1["B2"] = "='数据字典'!D7"
    ws1["C2"] = "='数据字典'!D8"

    # 下拉与校验
    dv_title = DataValidation(type="list", formula1="='数据字典'!$A$1:$A$6", allow_blank=True)
    dv_gender = DataValidation(type="list", formula1="='数据字典'!$C$1:$C$2", allow_blank=True)
    dv_date = DataValidation(type="date", operator="between",
                             formula1="DATE(1900,1,1)", formula2="DATE(9999,12,31)")
    ws1.add_data_validation(dv_title);  dv_title.add("D2")
    ws1.add_data_validation(dv_gender); dv_gender.add("E2")
    ws1.add_data_validation(dv_date);   dv_date.add("F2"); dv_date.add("G2")

    # 样式与列宽
    thin = Side(style="thin", color="CCCCCC")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for cell in ws1[2]:
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 8
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 12
    ws1.freeze_panes = "A2"
    ws1["F2"].number_format = "yyyy-mm-dd"
    ws1["G2"].number_format = "yyyy-mm-dd"

    # Sheet2：项目资料
    headers2 = ["申报日期", "课题名称", "课题编号", "金额", "开始日期", "结束日期", "状态"]
    ws2.append(headers2)

    # 预置空行
    for _ in range(2, max_rows + 1):
        ws2.append([None] * 7)

    dv_date2 = DataValidation(type="date", operator="between",
                              formula1="DATE(1900,1,1)", formula2="DATE(9999,12,31)")
    ws2.add_data_validation(dv_date2)
    dv_date2.add(f"A2:A{max_rows}")
    dv_date2.add(f"E2:E{max_rows}")
    dv_date2.add(f"F2:F{max_rows}")

    dv_amount = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0")
    ws2.add_data_validation(dv_amount)
    dv_amount.add(f"D2:D{max_rows}")

    dv_status = DataValidation(type="list", formula1="='数据字典'!$B$1:$B$4", allow_blank=True)
    ws2.add_data_validation(dv_status)
    dv_status.add(f"G2:G{max_rows}")

    # 样式与列宽
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws2.freeze_panes = "A2"
    widths = [12, 32, 18, 12, 12, 12, 10]
    for col, w in zip("ABCDEFG", widths):
        ws2.column_dimensions[col].width = w
    for row in ws2.iter_rows(min_row=2, max_row=max_rows, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r in range(2, max_rows + 1):
        ws2[f"A{r}"].number_format = "yyyy-mm-dd"
        ws2[f"E{r}"].number_format = "yyyy-mm-dd"
        ws2[f"F{r}"].number_format = "yyyy-mm-dd"
        ws2[f"D{r}"].number_format = "#,##0.00"

    # 隐藏数据字典
    #dict_ws.sheet_state = "hidden"

    wb.save(path)
    return path


if __name__ == "__main__":
    print("生成：", build_template())